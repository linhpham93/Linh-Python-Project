# How to read a column in excel
import openpyxl

wb = openpyxl.load_workbook("../Buoi6/data/example_file.xlsx")
sheet = wb.active

# Này là chỉ in ra cột chỉ định
for column in sheet.iter_cols(min_col=2, max_col=2, values_only=True):
    print(column)
    
# Này là in côt nhưng bắt đầu từ dòng thứ 2 
for  column2 in sheet.iter_cols(min_col=2, max_col=2, min_row=2, values_only=True): 
    print(column2)
    
# Này là in mà không cần dùng for
print(list(sheet.iter_rows(min_col=2, max_col=2, values_only=True)))